import os
import requests
from openpyxl import load_workbook

# Функция для получения прямой ссылки на изображение
def get_direct_link(yandex_disk_url):
    api_url = f"https://cloud-api.yandex.net/v1/disk/public/resources/download?public_key={yandex_disk_url}"
    response = requests.get(api_url)

    if response.status_code == 200:
        data = response.json()
        return data.get('href')  # Получаем ссылку на файл из тега "href"
    else:
        print(f"Ошибка получения прямой ссылки: {response.status_code}, {response.text}")
        return None

# Функция для создания папки и скачивания изображения
def process_excel_file(file_path):
    workbook = load_workbook(filename=file_path)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        folder_name = str(row[0])  # Название папки из первой ячейки
        image_url = row[1]          # URL изображения из второй ячейки

        # Создаем папку
        current_directory = os.getcwd()
        folder_path = os.path.join(current_directory, folder_name)

        if not os.path.exists(folder_path):
            os.makedirs(folder_path)

        # Скачиваем изображение
        if image_url:  # Проверяем, что URL изображения не пуст
            try:
                if image_url.startswith('https://disk.yandex.ru/'):
                    # Получаем прямую ссылку через API
                    direct_link = get_direct_link(image_url)

                    if direct_link:
                        # Скачиваем изображение по прямой ссылке из тега "href"
                        response = requests.get(direct_link)
                        response.raise_for_status()  # Проверяем статус ответа
                        if response.status_code == 200:
                            file_name = os.path.join(folder_path, f"{folder_name}.jpg")  # Используем имя папки для файла
                            with open(file_name, 'wb') as img_file:
                                img_file.write(response.content)
                            print(f"Изображение успешно скачано в {file_name}")
                        else:
                            print(f"Ошибка скачивания: статус {response.status_code} для URL: {direct_link}")
                    else:
                        print(f"Не удалось получить прямую ссылку для {image_url}")
                else:
                    # Прямой URL, не требующий обработки
                    response = requests.get(image_url)
                    response.raise_for_status()  # Проверяем статус ответа
                    if response.status_code == 200:
                        file_name = os.path.join(folder_path, f"{folder_name}.jpg")  # Используем имя папки для файла
                        with open(file_name, 'wb') as img_file:
                            img_file.write(response.content)
                        print(f"Изображение успешно скачано в {file_name}")
                    else:
                        print(f"Ошибка скачивания: статус {response.status_code} для URL: {image_url}")
            except requests.exceptions.RequestException as e:
                print(f"Ошибка скачивания изображения по URL {image_url}: {e}")

if __name__ == "__main__":
    process_excel_file('/Users/tsedilkin/Params/params.xlsx')
